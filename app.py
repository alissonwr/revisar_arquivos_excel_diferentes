import pandas as pd
import unidecode
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def normalizar_nomes_colunas(df):
    df.columns = [unidecode.unidecode(col.upper().strip()) for col in df.columns]
    return df

def encontrar_cidade_semelhante(cidade, lista_cidades, limiar=0.8):
    correspondencias = get_close_matches(cidade, lista_cidades, n=1, cutoff=limiar)
    return correspondencias[0] if correspondencias else None

def combinar_arquivos_excel(arquivo1, arquivo2, arquivo_saida):
    # Leitura dos arquivos
    df1 = pd.read_excel(arquivo1)
    df2 = pd.read_excel(arquivo2)
    
    # Normalização dos nomes das colunas
    df1 = normalizar_nomes_colunas(df1)
    df2 = normalizar_nomes_colunas(df2)
    
    # Normalização dos nomes das cidades
    df1['MUNICIPIO'] = df1['MUNICIPIO'].apply(lambda x: unidecode.unidecode(x.upper().strip()))
    df2['MUNICIPIO'] = df2['MUNICIPIO'].apply(lambda x: unidecode.unidecode(x.upper().strip()))
    
    # Encontrar cidades semelhantes e comuns
    df1['MUNICIPIO_ARQUIVO2'] = df1['MUNICIPIO'].apply(lambda x: encontrar_cidade_semelhante(x, df2['MUNICIPIO'].tolist()))
    df1 = df1.dropna(subset=['MUNICIPIO_ARQUIVO2'])
    
    df2_comum = df2[df2['MUNICIPIO'].isin(df1['MUNICIPIO_ARQUIVO2'])]
    
    # Combinar os dataframes
    df_combinado = pd.merge(df1, df2_comum, left_on='MUNICIPIO_ARQUIVO2', right_on='MUNICIPIO', suffixes=('_ARQUIVO1', '_ARQUIVO2'))
    
    # Reordenar colunas para colocar as semelhantes lado a lado
    colunas_arquivo1 = [col for col in df_combinado.columns if '_ARQUIVO1' in col]
    colunas_arquivo2 = [col for col in df_combinado.columns if '_ARQUIVO2' in col]
    
    colunas_ordenadas = []
    for col in colunas_arquivo1:
        base_coluna = col.replace('_ARQUIVO1', '')
        if f'{base_coluna}_ARQUIVO2' in colunas_arquivo2:
            colunas_ordenadas.append(col)
            colunas_ordenadas.append(f'{base_coluna}_ARQUIVO2')
    
    # Adicionar colunas que não têm correspondência
    colunas_restantes = [col for col in df_combinado.columns if col not in colunas_ordenadas]
    colunas_ordenadas.extend(colunas_restantes)
    
    df_combinado = df_combinado[colunas_ordenadas]
    
    # Salvar o resultado em um arquivo Excel
    df_combinado.to_excel(arquivo_saida, index=False)
    
    # Aplicar a formatação condicional para células diferentes
    destacar_diferencas(arquivo_saida)
    
    return df_combinado

def destacar_diferencas(arquivo_saida):
    # Carregar o arquivo Excel
    wb = load_workbook(arquivo_saida)
    ws = wb.active
    
    # Definir o preenchimento para as células com diferenças
    preenchimento_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Obter o número de colunas e linhas
    numero_colunas = ws.max_column
    numero_linhas = ws.max_row
    
    # Iterar sobre as colunas e comparar os valores
    for col in range(1, numero_colunas, 2):  # Considera pares de colunas
        nome_coluna1 = ws.cell(row=1, column=col).value
        nome_coluna2 = ws.cell(row=1, column=col + 1).value

        # Verificar se os nomes das colunas são idênticos (removendo os sufixos "_ARQUIVO1" e "_ARQUIVO2")
        if nome_coluna1 and nome_coluna2 and nome_coluna1.split('_ARQUIVO')[0] == nome_coluna2.split('_ARQUIVO')[0]:
            for linha in range(2, numero_linhas + 1):
                celula1 = ws.cell(row=linha, column=col)
                celula2 = ws.cell(row=linha, column=col + 1)
                
                # Comparar valores e aplicar formatação se diferentes
                if celula1.value != celula2.value:
                    celula1.fill = preenchimento_amarelo
                    celula2.fill = preenchimento_amarelo
    
    # Salvar o arquivo com as diferenças destacadas
    wb.save(arquivo_saida)

# Flask para a interface web
from flask import Flask, request, render_template, send_file

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        arquivo1 = request.files['arquivo1']
        arquivo2 = request.files['arquivo2']
        arquivo_saida = "saida_combinada.xlsx"
        
        # Salvar os arquivos no servidor
        caminho_arquivo1 = f"./{arquivo1.filename}"
        caminho_arquivo2 = f"./{arquivo2.filename}"
        arquivo1.save(caminho_arquivo1)
        arquivo2.save(caminho_arquivo2)
        
        # Fazer a combinação dos arquivos
        combinar_arquivos_excel(caminho_arquivo1, caminho_arquivo2, arquivo_saida)
        
        # Retornar o arquivo Excel resultante para download
        return send_file(arquivo_saida, as_attachment=True)
    
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
